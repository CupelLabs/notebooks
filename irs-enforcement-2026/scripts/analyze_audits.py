"""
IRS Enforcement Collapse Analysis
Cupel Labs — 2026-03-21

Analyzes IRS Data Book tables to show:
1. Audit rate trends by income bracket (2010-2022)
2. IRS staffing collapse (1995-2024)
3. The inverted audit curve (poor vs rich)
4. Revenue efficiency (collections per FTE)
"""

import openpyxl
import pandas as pd
import json
import os

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(BASE, "data", "raw")
INTERMEDIATE = os.path.join(BASE, "data", "intermediate")
os.makedirs(INTERMEDIATE, exist_ok=True)

# ============================================================
# 1. Parse Table 17: Audit rates by income bracket over time
# ============================================================

print("=" * 60)
print("PARSING TABLE 17: Audit Coverage by Income Bracket")
print("=" * 60)

# Income bracket rows we care about (individual returns)
BRACKET_LABELS = {
	"$1 under $25,000": "$1-$25K",
	"$25,000 under $50,000": "$25K-$50K",
	"$50,000 under $75,000": "$50K-$75K",
	"$75,000 under $100,000": "$75K-$100K",
	"$100,000 under $200,000": "$100K-$200K",
	"$200,000 under $500,000": "$200K-$500K",
	"$500,000 under $1,000,000": "$500K-$1M",
	"$1,000,000 under $5,000,000": "$1M-$5M",
	"$5,000,000 under $10,000,000": "$5M-$10M",
	"$10,000,000 or more": "$10M+",
}

EITC_LABEL = "Returns with earned income tax credit"
INDIVIDUAL_TOTAL = "Individual income tax returns, total"

def parse_table17(filepath):
	"""Parse a Table 17 Excel file and extract audit rates by income bracket."""
	wb = openpyxl.load_workbook(filepath, data_only=True)
	ws = wb[wb.sheetnames[0]]

	# Find tax year columns by scanning row 3
	tax_year_cols = {}
	for cell in ws[3]:
		val = str(cell.value) if cell.value else ""
		if "Tax Year" in val:
			# Extract year
			year = None
			for word in val.split():
				try:
					y = int(word.replace("†", "").replace("[", "").replace("]", ""))
					if 2000 <= y <= 2030:
						year = y
						break
				except ValueError:
					continue
			if year:
				tax_year_cols[year] = cell.column

	print(f"  Found tax years: {sorted(tax_year_cols.keys())}")

	# For each tax year column, find:
	# - Column offset +0: returns filed
	# - Column offset +3: percentage covered (audit rate)
	# - Column offset +5: recommended additional tax

	results = []
	for year, base_col in sorted(tax_year_cols.items()):
		# Scan rows for income brackets
		for row_idx in range(7, ws.max_row + 1):
			label_cell = ws.cell(row=row_idx, column=1).value
			if not label_cell:
				continue
			label = str(label_cell).strip()

			# Check if this row is a bracket we want
			matched_bracket = None
			for pattern, short in BRACKET_LABELS.items():
				if label.startswith(pattern):
					matched_bracket = short
					break

			is_eitc = label.startswith(EITC_LABEL)
			is_individual_total = label.startswith(INDIVIDUAL_TOTAL)

			if matched_bracket or is_eitc or is_individual_total:
				returns_filed = ws.cell(row=row_idx, column=base_col).value
				pct_covered = ws.cell(row=row_idx, column=base_col + 3).value
				rec_add_tax = ws.cell(row=row_idx, column=base_col + 5).value

				# Clean values
				try:
					returns_filed = int(returns_filed) if returns_filed and str(returns_filed) not in ("[6]", "d", "") else None
				except (ValueError, TypeError):
					returns_filed = None
				try:
					pct_covered = float(pct_covered) if pct_covered and str(pct_covered) not in ("[6]", "d", "") else None
				except (ValueError, TypeError):
					pct_covered = None
				try:
					rec_add_tax = float(rec_add_tax) if rec_add_tax and str(rec_add_tax) not in ("[6]", "d", "") else None
				except (ValueError, TypeError):
					rec_add_tax = None

				category = "EITC" if is_eitc else ("Individual Total" if is_individual_total else matched_bracket)

				results.append({
					"tax_year": year,
					"bracket": category,
					"returns_filed": returns_filed,
					"audit_rate_pct": pct_covered,
					"rec_additional_tax_thousands": rec_add_tax,
				})

	return results


# Parse both primary Table 17 files
print("\nParsing revised all-years file (2010-2021)...")
results_main = parse_table17(os.path.join(RAW, "table17_exam_coverage_all_years_2010_2021_revised.xlsx"))

print("\nParsing 2024 Data Book file (adds 2022)...")
results_2024 = parse_table17(os.path.join(RAW, "table17_exam_coverage_2024db.xlsx"))

# Merge: use 2024db for 2022, main file for 2010-2021
all_results = results_main.copy()
years_in_main = {r["tax_year"] for r in results_main}
for r in results_2024:
	if r["tax_year"] not in years_in_main:
		all_results.append(r)

df_audits = pd.DataFrame(all_results)
df_audits = df_audits.sort_values(["tax_year", "bracket"]).reset_index(drop=True)

print(f"\nTotal records: {len(df_audits)}")
print(f"Tax years: {sorted(df_audits['tax_year'].unique())}")
print(f"Brackets: {df_audits['bracket'].unique()}")

# Export
df_audits.to_csv(os.path.join(INTERMEDIATE, "audit_rates_by_bracket_2010_2022.csv"), index=False)
print(f"\nExported to {INTERMEDIATE}/audit_rates_by_bracket_2010_2022.csv")


# ============================================================
# 2. Parse Table 33: Staffing and Collections (1995-2024)
# ============================================================

print("\n" + "=" * 60)
print("PARSING TABLE 33: Staffing, Collections, Costs")
print("=" * 60)

wb33 = openpyxl.load_workbook(os.path.join(RAW, "table33_collections_costs_personnel_1995_2024.xlsx"), data_only=True)
ws33 = wb33[wb33.sheetnames[0]]

staffing_data = []
for row_idx in range(4, 34):  # Rows 4-33 are 1995-2024
	year_val = ws33.cell(row=row_idx, column=1).value
	if not year_val:
		continue
	try:
		year = int(year_val)
	except (ValueError, TypeError):
		continue

	gross_collections = ws33.cell(row=row_idx, column=2).value
	operating_costs = ws33.cell(row=row_idx, column=3).value
	cost_per_100 = ws33.cell(row=row_idx, column=4).value
	population = ws33.cell(row=row_idx, column=5).value
	per_capita = ws33.cell(row=row_idx, column=6).value
	fte = ws33.cell(row=row_idx, column=7).value

	staffing_data.append({
		"fiscal_year": year,
		"gross_collections_thousands": int(gross_collections) if gross_collections else None,
		"operating_costs_thousands": int(operating_costs) if operating_costs else None,
		"cost_per_100_collected": float(cost_per_100) if cost_per_100 else None,
		"us_population_thousands": int(population) if population else None,
		"per_capita_tax": float(per_capita) if per_capita else None,
		"fte_positions": int(fte) if fte else None,
	})

df_staff = pd.DataFrame(staffing_data)
df_staff.to_csv(os.path.join(INTERMEDIATE, "irs_staffing_1995_2024.csv"), index=False)
print(f"Exported {len(df_staff)} years to {INTERMEDIATE}/irs_staffing_1995_2024.csv")


# ============================================================
# 3. ANALYSIS AND FINDINGS
# ============================================================

print("\n" + "=" * 60)
print("ANALYSIS: Key Findings")
print("=" * 60)

# --- Finding 1: The audit rate collapse for high-income filers ---
print("\n--- Finding 1: Audit Rate Collapse by Income Bracket ---")
brackets_of_interest = ["$1-$25K", "$200K-$500K", "$1M-$5M", "$5M-$10M", "$10M+", "Individual Total", "EITC"]
for bracket in brackets_of_interest:
	bdata = df_audits[df_audits["bracket"] == bracket].sort_values("tax_year")
	if len(bdata) == 0:
		continue
	first = bdata[bdata["audit_rate_pct"].notna()].iloc[0] if len(bdata[bdata["audit_rate_pct"].notna()]) > 0 else None
	last = bdata[bdata["audit_rate_pct"].notna()].iloc[-1] if len(bdata[bdata["audit_rate_pct"].notna()]) > 0 else None
	if first is not None and last is not None:
		change = ((last["audit_rate_pct"] - first["audit_rate_pct"]) / first["audit_rate_pct"]) * 100
		print(f"  {bracket:>20s}: {first['audit_rate_pct']:6.3f}% ({int(first['tax_year'])}) → {last['audit_rate_pct']:6.3f}% ({int(last['tax_year'])}) | Change: {change:+.1f}%")

# --- Finding 2: The inverted audit curve (2022) ---
print("\n--- Finding 2: The Inverted Audit Curve (Latest Year) ---")
latest_year = df_audits["tax_year"].max()
latest = df_audits[(df_audits["tax_year"] == latest_year) & (df_audits["bracket"].isin(BRACKET_LABELS.values()))].copy()
latest = latest.sort_values("audit_rate_pct", ascending=False)
print(f"  Tax Year {int(latest_year)} audit rates:")
for _, row in latest.iterrows():
	if row["audit_rate_pct"] is not None:
		print(f"    {row['bracket']:>15s}: {row['audit_rate_pct']:.3f}%")

# Find the EITC vs upper-middle comparison
eitc_rate = df_audits[(df_audits["tax_year"] == latest_year) & (df_audits["bracket"] == "EITC")]["audit_rate_pct"].values
low_income = df_audits[(df_audits["tax_year"] == latest_year) & (df_audits["bracket"] == "$1-$25K")]["audit_rate_pct"].values
upper_middle = df_audits[(df_audits["tax_year"] == latest_year) & (df_audits["bracket"] == "$200K-$500K")]["audit_rate_pct"].values

if len(low_income) > 0 and len(upper_middle) > 0 and upper_middle[0] and upper_middle[0] > 0:
	ratio = low_income[0] / upper_middle[0]
	print(f"\n  Low-income ($1-$25K) audit rate is {ratio:.1f}x the $200K-$500K rate")
if len(eitc_rate) > 0 and len(upper_middle) > 0 and upper_middle[0] and upper_middle[0] > 0:
	ratio = eitc_rate[0] / upper_middle[0]
	print(f"  EITC audit rate is {ratio:.1f}x the $200K-$500K rate")

# --- Finding 3: Staffing collapse ---
print("\n--- Finding 3: IRS Staffing Collapse ---")
peak = df_staff.loc[df_staff["fte_positions"].idxmax()]
trough = df_staff.loc[df_staff["fte_positions"].idxmin()]
current = df_staff.iloc[-1]
print(f"  Peak: {int(peak['fte_positions']):,} FTEs ({int(peak['fiscal_year'])})")
print(f"  Trough: {int(trough['fte_positions']):,} FTEs ({int(trough['fiscal_year'])})")
print(f"  Current: {int(current['fte_positions']):,} FTEs ({int(current['fiscal_year'])})")
print(f"  Peak-to-trough decline: {(trough['fte_positions'] - peak['fte_positions']) / peak['fte_positions'] * 100:.1f}%")
print(f"  Still below peak by: {(current['fte_positions'] - peak['fte_positions']) / peak['fte_positions'] * 100:.1f}%")

# --- Finding 4: Revenue efficiency ---
print("\n--- Finding 4: Revenue Efficiency ---")
# Collections per FTE
df_staff["collections_per_fte"] = df_staff["gross_collections_thousands"] * 1000 / df_staff["fte_positions"]
first_yr = df_staff.iloc[0]
last_yr = df_staff.iloc[-1]
print(f"  Collections per FTE ({int(first_yr['fiscal_year'])}): ${first_yr['collections_per_fte']:,.0f}")
print(f"  Collections per FTE ({int(last_yr['fiscal_year'])}): ${last_yr['collections_per_fte']:,.0f}")
print(f"  Change: {(last_yr['collections_per_fte'] / first_yr['collections_per_fte'] - 1) * 100:.0f}%")

# --- Finding 5: Recommended additional tax from high-income audits ---
print("\n--- Finding 5: Additional Tax from High-Income Audits ---")
high_income_brackets = ["$1M-$5M", "$5M-$10M", "$10M+"]
for year in [2010, 2015, 2019, 2021, 2022]:
	year_data = df_audits[(df_audits["tax_year"] == year) & (df_audits["bracket"].isin(high_income_brackets))]
	total_add_tax = year_data["rec_additional_tax_thousands"].sum()
	if total_add_tax > 0:
		print(f"  {year}: ${total_add_tax / 1e6:.2f}B in recommended additional tax from $1M+ audits")

# --- Finding 6: The $10M+ audit rate time series ---
print("\n--- Finding 6: $10M+ Audit Rate Over Time ---")
ten_m = df_audits[df_audits["bracket"] == "$10M+"].sort_values("tax_year")
for _, row in ten_m.iterrows():
	if row["audit_rate_pct"] is not None:
		print(f"  {int(row['tax_year'])}: {row['audit_rate_pct']:.2f}%")

# --- Finding 7: Total returns filed vs examined ---
print("\n--- Finding 7: Total Individual Returns Filed vs Examined ---")
individual = df_audits[df_audits["bracket"] == "Individual Total"].sort_values("tax_year")
for _, row in individual.iterrows():
	if row["audit_rate_pct"] is not None:
		print(f"  {int(row['tax_year'])}: {row['returns_filed']:>12,} filed, audit rate {row['audit_rate_pct']:.3f}%")

# --- Finding 8: The EITC targeting ---
print("\n--- Finding 8: EITC Audit Rates Over Time ---")
eitc = df_audits[df_audits["bracket"] == "EITC"].sort_values("tax_year")
for _, row in eitc.iterrows():
	if row["audit_rate_pct"] is not None:
		print(f"  {int(row['tax_year'])}: {row['audit_rate_pct']:.3f}%")

# Export summary for charts
print("\n" + "=" * 60)
print("EXPORTING CHART DATA")
print("=" * 60)

# Chart data 1: Audit rates for key brackets over time
chart_brackets = ["$1-$25K", "$200K-$500K", "$1M-$5M", "$10M+", "EITC"]
chart_data = df_audits[df_audits["bracket"].isin(chart_brackets)][["tax_year", "bracket", "audit_rate_pct"]].copy()
chart_data = chart_data.dropna(subset=["audit_rate_pct"])
# Deduplicate: keep first (from revised all-years file, which is more authoritative)
chart_data = chart_data.drop_duplicates(subset=["tax_year", "bracket"], keep="first")
pivot = chart_data.pivot(index="tax_year", columns="bracket", values="audit_rate_pct")
pivot.to_csv(os.path.join(INTERMEDIATE, "audit_rates_key_brackets_pivot.csv"))
print("Exported audit_rates_key_brackets_pivot.csv")

# Chart data 2: Staffing time series
df_staff[["fiscal_year", "fte_positions", "gross_collections_thousands", "collections_per_fte"]].to_csv(
	os.path.join(INTERMEDIATE, "irs_staffing_efficiency.csv"), index=False
)
print("Exported irs_staffing_efficiency.csv")

# Chart data 3: Latest year income bracket comparison (the inverted curve)
latest_brackets = df_audits[(df_audits["tax_year"] == latest_year) & (df_audits["bracket"].isin(BRACKET_LABELS.values()))].copy()
latest_brackets = latest_brackets.drop_duplicates(subset=["bracket"], keep="first")
latest_brackets = latest_brackets[["bracket", "audit_rate_pct", "returns_filed"]].dropna()
latest_brackets.to_csv(os.path.join(INTERMEDIATE, "audit_rates_latest_by_bracket.csv"), index=False)
print("Exported audit_rates_latest_by_bracket.csv")

# JSON for React chart components
chart_json = {
	"audit_rates_over_time": [],
	"staffing": [],
	"latest_brackets": [],
}

for _, row in chart_data.iterrows():
	chart_json["audit_rates_over_time"].append({
		"year": int(row["tax_year"]),
		"bracket": row["bracket"],
		"rate": round(row["audit_rate_pct"], 3),
	})

for _, row in df_staff.iterrows():
	chart_json["staffing"].append({
		"year": int(row["fiscal_year"]),
		"fte": int(row["fte_positions"]),
		"collections_billions": round(row["gross_collections_thousands"] / 1e6, 1),
	})

for _, row in latest_brackets.iterrows():
	chart_json["latest_brackets"].append({
		"bracket": row["bracket"],
		"rate": round(row["audit_rate_pct"], 3),
		"returns_filed": int(row["returns_filed"]) if pd.notna(row["returns_filed"]) else None,
	})

with open(os.path.join(INTERMEDIATE, "chart_data.json"), "w") as f:
	json.dump(chart_json, f, indent=2)
print("Exported chart_data.json")

print("\nDone.")
